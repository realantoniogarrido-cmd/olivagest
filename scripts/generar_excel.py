#!/usr/bin/env python3
"""
Genera un Excel estético para OlivaGest.
Recibe JSON por stdin: { "tipo": "...", "hoja": "...", "columnas": [...], "filas": [...] }
Devuelve el archivo .xlsx por stdout (bytes).
"""
import sys, json, io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Paleta OlivaGest ────────────────────────────────────────────────
NAVY        = "0F172A"
NAVY_LIGHT  = "1E293B"
GREEN       = "16A34A"
GREEN_LIGHT = "DCFCE7"
GRAY_BG     = "F8FAFC"
GRAY_LINE   = "E2E8F0"
WHITE       = "FFFFFF"
AMBER       = "D97706"
AMBER_LIGHT = "FFFBEB"

def make_border(color=GRAY_LINE, style="thin"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def make_bottom_border(color=GRAY_LINE):
    return Border(bottom=Side(border_style="thin", color=color))

def col_width(values, header, min_w=10, max_w=40):
    longest = max((len(str(v)) for v in values if v is not None), default=0)
    return min(max(longest, len(header), min_w), max_w)

def generar(tipo, hoja_nombre, columnas, filas):
    wb = Workbook()
    ws = wb.active
    ws.title = hoja_nombre

    n_cols = len(columnas)

    # ── Fila 1: Banner OlivaGest ────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    banner = ws.cell(1, 1)
    banner.value = f"OlivaGest · {hoja_nombre}"
    banner.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    banner.fill = PatternFill("solid", fgColor=NAVY)
    banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # ── Fila 2: Subtítulo / fecha ────────────────────────────────────
    from datetime import date
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub = ws.cell(2, 1)
    sub.value = f"Exportado el {date.today().strftime('%d/%m/%Y')}  ·  {len(filas)} registros"
    sub.font = Font(name="Arial", size=9, italic=True, color="94A3B8")
    sub.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # ── Fila 3: Encabezados ──────────────────────────────────────────
    HEADER_ROW = 3
    for c, col in enumerate(columnas, 1):
        cell = ws.cell(HEADER_ROW, c)
        cell.value = col["label"]
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor="16A34A")  # verde cooperativa
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border(GREEN, "thin")
    ws.row_dimensions[HEADER_ROW].height = 22

    # ── Datos ────────────────────────────────────────────────────────
    for r, fila in enumerate(filas, HEADER_ROW + 1):
        bg = WHITE if (r - HEADER_ROW) % 2 == 1 else GRAY_BG
        for c, col in enumerate(columnas, 1):
            val = fila.get(col["key"], "")
            cell = ws.cell(r, c)

            # Formateo según tipo de columna
            if col.get("tipo") == "numero" and val not in (None, ""):
                try:
                    cell.value = float(val)
                    cell.number_format = col.get("formato", "#,##0.00")
                except:
                    cell.value = val
            elif col.get("tipo") == "euro" and val not in (None, ""):
                try:
                    cell.value = float(val)
                    cell.number_format = '#,##0.00 "€"'
                except:
                    cell.value = val
            elif col.get("tipo") == "kg" and val not in (None, ""):
                try:
                    cell.value = float(val)
                    cell.number_format = '#,##0.00 "kg"'
                except:
                    cell.value = val
            elif col.get("tipo") == "pct" and val not in (None, ""):
                try:
                    cell.value = float(val)
                    cell.number_format = '0.00"%"'
                except:
                    cell.value = val
            else:
                cell.value = val if val not in (None, "") else "—"

            cell.font = Font(name="Arial", size=10, color="1E293B")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal=col.get("align", "left"),
                vertical="center"
            )
            cell.border = make_bottom_border(GRAY_LINE)

        ws.row_dimensions[r].height = 18

    # ── Fila de totales (si hay columnas de suma) ────────────────────
    cols_suma = [c for c in columnas if c.get("suma")]
    if cols_suma and filas:
        tot_row = HEADER_ROW + len(filas) + 1
        for c, col in enumerate(columnas, 1):
            cell = ws.cell(tot_row, c)
            if c == 1:
                cell.value = "TOTAL"
                cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            elif col.get("suma"):
                data_start = HEADER_ROW + 1
                data_end   = HEADER_ROW + len(filas)
                col_letter = get_column_letter(c)
                cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
                fmt = col.get("formato") or ('#,##0.00 "€"' if col.get("tipo") == "euro" else '#,##0.00 "kg"')
                cell.number_format = fmt
                cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.fill = PatternFill("solid", fgColor=NAVY)
        ws.row_dimensions[tot_row].height = 22

    # ── Anchos de columna ────────────────────────────────────────────
    for c, col in enumerate(columnas, 1):
        vals = [fila.get(col["key"], "") for fila in filas]
        ws.column_dimensions[get_column_letter(c)].width = col_width(vals, col["label"],
            min_w=col.get("min_w", 10), max_w=col.get("max_w", 40))

    # ── Freeze panes (fijar encabezado) ─────────────────────────────
    ws.freeze_panes = f"A{HEADER_ROW + 1}"

    # ── Autofilter ───────────────────────────────────────────────────
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(n_cols)}{HEADER_ROW + len(filas)}"

    # ── Salida ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    payload = json.loads(sys.stdin.read())
    resultado = generar(
        payload["tipo"],
        payload["hoja"],
        payload["columnas"],
        payload["filas"],
    )
    sys.stdout.buffer.write(resultado)
